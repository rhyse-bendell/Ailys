import sys
import os
import traceback
import re
import openpyxl
from dotenv import load_dotenv
from openai import OpenAI
from core.pdf_reader import extract_text_from_pdf
from openpyxl.utils import get_column_letter
from memory.memory import save_memory_event, get_ai_insights_by_tag
from core.approval_queue import request_approval
# Set up environment and OpenAI client
load_dotenv()

_client = None
def get_openai_client():
    global _client
    if _client is None:

        key = os.getenv("OPENAI_API_KEY")
        if not key:
            raise ValueError("OPENAI_API_KEY is not set")
        _client = OpenAI(api_key=key)
    return _client


COLUMNS = [
    "Source", "Citation", "Brief Summary", "Abstract", "Expanded Notes",
    "Research Questions", "Hypotheses", "Referenced works to review",
    "Keywords", "Type of Publication", "Reviewed By"
]

def get_memory_context(tag="literature_review", max_memories=5):
    insights = get_ai_insights_by_tag(tag)
    return "\n\n".join(insights[-int(max_memories):]) if insights else ""

FIELD_PROMPT = """
You are Ailys, a research assistant trained to produce structured literature reviews.
Given the text of a scholarly academic article, extract the following fields **completely and with exactitude**:

Instructions for each field:
- **Abstract**: Extract the full, verbatim abstract from the article. Do not paraphrase. Enclose in quotation marks.
- **Expanded Notes**: Provide 30–50 direct quotes from the article, each one bolding key terms. Include page numbers if possible.
- **Referenced works to review**: Find and list full APA-style references from the article’s References section that are important or highly relevant.
- All other fields should be concise but accurate, using direct text where possible.
- Do not skip any fields. If a field is missing, write "None".

Previous Relevant Memory Insights:
{memory_context}

Return your answer in this format:
Source: <insert>
Citation: <insert>
Brief Summary: <insert>
Abstract: "<verbatim abstract>"
Expanded Notes: "<30–50 quoted entries with bolded terms>"
Research Questions: <insert>
Hypotheses: <insert>
Referenced works to review: <insert full APA references>
Keywords: <insert>
Type of Publication: <insert>
Reviewed By: Rhyse + AI Assistant

Guidance for this specific review: {guidance}

Begin processing the following article text:
{text}
"""

def run(pdf_path, guidance="", recall_depth=5, output_file="outputs/C2_Lit_Review.xlsx"):
    print(f"📄 Starting literature review for: {os.path.basename(pdf_path)}")
    recall_depth = int(recall_depth)

    try:
        full_text, error = extract_text_from_pdf(pdf_path)
        if error:
            print(f"❌ PDF extraction error: {error}")
            return False, f"PDF extraction failed: {error}"
        if not isinstance(full_text, str) or not full_text.strip():
            print("❌ PDF text is empty or invalid. Skipping review.")
            return False, "PDF text is empty or could not be parsed."
    except Exception as e:
        print("❌ Exception during PDF extraction:")
        traceback.print_exc()
        return False, f"Exception during PDF extraction: {e}"

    try:
        text_chunk = full_text[:10000]
        memory_context = get_memory_context(max_memories=recall_depth)

        prompt = FIELD_PROMPT.format(text=text_chunk, guidance=guidance, memory_context=memory_context)

        print("🧠 Sending prompt to GPT-4 for review extraction...")
        response = request_approval(
            description=f"Run GPT-4 review on '{os.path.basename(pdf_path)}'",
            call_fn=lambda: get_openai_client().chat.completions.create(
                model="gpt-4",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.3
            )
        )

        if not response or not hasattr(response, "choices") or not response.choices:
            print("❌ No valid response returned from GPT.")
            return False, "GPT did not return a valid response."

        content = response.choices[0].message.content
        print("✅ Received response from GPT-4.")
    except Exception as e:
        print("❌ Exception during GPT prompt call:")
        traceback.print_exc()
        return False, f"GPT prompt failed: {e}"

    try:
        parsed_data = parse_fields(content)
    except Exception as e:
        print("❌ Exception while parsing GPT output:")
        traceback.print_exc()
        return False, f"Failed to parse GPT output: {e}"

    try:
        append_to_spreadsheet(parsed_data, output_file)
    except Exception as e:
        print("❌ Exception while writing to spreadsheet:")
        traceback.print_exc()
        return False, f"Failed to write to spreadsheet: {e}"

    memory_summary_prompt = f"""
You are Ailys, a memory encoding system. Your task is to extract and compress the most important structured insights from a literature review task into a short but information-rich memory trace.
Focus on what this review adds to your understanding of team cognition, macrocognition, mental models, and how reviews are typically structured.
Respond in exactly 3-5 sentences, in plain declarative English.

Review Fields:
{content}
"""

    try:
        print("🧠 Sending insight prompt to GPT-4...")
        memory_response = get_openai_client().chat.completions.create(
            model="gpt-4",
            messages=[{"role": "user", "content": memory_summary_prompt}],
            temperature=0.3
        )
        ai_insight = memory_response.choices[0].message.content.strip()
    except Exception as e:
        print("❌ Failed to generate insight summary:")
        traceback.print_exc()
        ai_insight = "Insight not generated due to error."

    try:
        output_path = os.path.abspath(output_file)
        save_memory_event(
            event_type="literature_review",
            source_text=content,
            ai_insight=ai_insight,
            user_input=guidance,
            tags=["literature_review", "structured", "PDF"],
            file_path=output_path
        )
    except Exception as e:
        print("❌ Failed to save memory:")
        traceback.print_exc()
        return False, f"Failed to save memory event: {e}"

    print("✅ Review complete and saved.\n")
    return True, f"Review completed and saved to: {output_path}"



def parse_fields(text):
    parsed = []
    for field in COLUMNS:
        match = re.search(rf"{re.escape(field)}:\s*(.*?)(?=\n[A-Z][a-zA-Z ]+?:|$)", text, re.DOTALL)
        value = match.group(1).strip().replace("\n", " ").replace("  ", " ") if match else ""
        if field == "Reviewed By" and not value:
            value = "Rhyse + AI Assistant"
        parsed.append(value)
    return parsed

def append_to_spreadsheet(data, file_path):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(COLUMNS)
    else:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

    data = (data + [""] * len(COLUMNS))[:len(COLUMNS)]
    ws.append(data)
    wb.save(file_path)
